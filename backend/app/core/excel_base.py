"""
Excel Base - Excel处理基类
"""
import pandas as pd
import openpyxl
from pathlib import Path
from typing import Optional
import time


class ExcelBase:
    """统一Excel读写，处理权限错误重试"""

    def safe_read_excel(self, path: str, **kwargs) -> pd.DataFrame:
        """
        安全读取Excel文件

        Args:
            path: 文件路径
            **kwargs: 传递给pd.read_excel的参数

        Returns:
            DataFrame
        """
        max_retries = 3
        for attempt in range(max_retries):
            try:
                return pd.read_excel(path, **kwargs)
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(1)
                    continue
                raise
            except Exception as e:
                raise e

    def safe_load_workbook(self, path: str, **kwargs):
        """
        安全加载Excel工作簿

        Args:
            path: 文件路径
            **kwargs: 传递给openpyxl.load_workbook的参数

        Returns:
            Workbook对象
        """
        max_retries = 3
        for attempt in range(max_retries):
            try:
                return openpyxl.load_workbook(path, **kwargs)
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(1)
                    continue
                raise
            except Exception as e:
                raise e

    def safe_write_excel(self, df: pd.DataFrame, path: str, **kwargs):
        """
        安全写入Excel文件

        Args:
            df: DataFrame
            path: 输出路径
            **kwargs: 传递给df.to_excel的参数
        """
        max_retries = 3
        for attempt in range(max_retries):
            try:
                df.to_excel(path, **kwargs)
                return
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(1)
                    continue
                raise
            except Exception as e:
                raise e

    def safe_save_workbook(self, wb, path: str):
        """
        安全保存工作簿

        Args:
            wb: Workbook对象
            path: 输出路径
        """
        max_retries = 3
        for attempt in range(max_retries):
            try:
                wb.save(path)
                return
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(1)
                    continue
                raise
            except Exception as e:
                raise e